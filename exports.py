"""
exports.py
----------
Builds a formatted Excel workbook from a variance report.

Why a module rather than a `df.to_excel()` call in the dashboard: FP&A
recipients live in Excel, and a raw dump is not a deliverable. This
produces something an analyst could forward without touching it —
number formats, frozen headers, autofilters, sensible column widths, and
a summary tab that answers "what happened this period" before anyone
scrolls.

Three sheets:
  Summary      -- run parameters, headline figures, pattern counts, and a
                  department roll-up
  Flagged      -- every breaching line with severity, pattern, commentary
  Detail       -- the full filtered dataset behind the numbers

Kept free of Streamlit imports so it can be tested headlessly and reused
by a scheduled job.
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["build_excel_report", "template_csv_bytes", "TEMPLATE_COLUMNS"]

# --- house style -----------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
LABEL_FONT = Font(bold=True, size=10)
BAD_FILL = PatternFill("solid", fgColor="FCE4E4")
GOOD_FILL = PatternFill("solid", fgColor="E4F3E4")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = '#,##0;[Red](#,##0)'
PCT_FMT = '0.0%;[Red](0.0%)'

TEMPLATE_COLUMNS = [
    "month", "department", "account", "account_type", "budget", "actual",
]


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _write_table(
    ws: Worksheet,
    df: pd.DataFrame,
    start_row: int = 1,
    money_cols: tuple[str, ...] = (),
    pct_cols: tuple[str, ...] = (),
    widths: dict[str, int] | None = None,
    highlight_col: str | None = None,
) -> int:
    """Write a dataframe as a formatted table. Returns the last row used."""
    widths = widths or {}

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            value = row[col]
            if pd.isna(value):
                value = None
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = BORDER
            if col in money_cols:
                cell.number_format = MONEY_FMT
            elif col in pct_cols:
                cell.number_format = PCT_FMT
            if col == "commentary":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        if highlight_col and highlight_col in df.columns:
            v = row[highlight_col]
            if pd.notna(v) and v != 0:
                fill = GOOD_FILL if v > 0 else BAD_FILL
                for j in range(1, len(df.columns) + 1):
                    ws.cell(row=i, column=j).fill = fill

    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        if col in widths:
            ws.column_dimensions[letter].width = widths[col]
        else:
            longest = max(
                [len(str(col))] + [len(str(v)) for v in df[col].head(200)]
            )
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 40)

    last_row = start_row + len(df)
    if len(df):
        ws.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(df.columns))}{last_row}"
        )
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return last_row


def _kv(ws: Worksheet, row: int, label: str, value, fmt: str | None = None) -> int:
    """Write a label/value pair on the summary sheet."""
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    c = ws.cell(row=row, column=2, value=value)
    if fmt:
        c.number_format = fmt
    return row + 1


# ---------------------------------------------------------------------------
# main entry point
# ---------------------------------------------------------------------------
def build_excel_report(
    detail: pd.DataFrame,
    flagged: pd.DataFrame,
    view_label: str = "Operating expenses",
    pct_threshold: float = 0.10,
    dollar_floor: float = 10_000,
) -> bytes:
    """Assemble the workbook and return it as bytes, ready for download.

    `detail` is the filtered variance report; `flagged` is the breaching
    subset with commentary already attached.
    """
    from openpyxl import Workbook

    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Budget vs Actual — Variance Summary"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    r = 3
    r = _kv(ws, r, "Generated", datetime.now().strftime("%d %b %Y %H:%M"))
    r = _kv(ws, r, "View", view_label)
    if len(detail):
        r = _kv(ws, r, "Period",
                f"{detail['month'].min()} to {detail['month'].max()}")
        r = _kv(ws, r, "Departments", ", ".join(sorted(detail["department"].unique())))
    r = _kv(ws, r, "Variance threshold", pct_threshold, PCT_FMT)
    r = _kv(ws, r, "Materiality floor", dollar_floor, MONEY_FMT)

    r += 1
    ws.cell(row=r, column=1, value="Headline").font = TITLE_FONT
    r += 1
    total_budget = float(detail["budget"].sum()) if len(detail) else 0.0
    total_actual = float(detail["actual"].sum()) if len(detail) else 0.0
    variance = total_actual - total_budget
    r = _kv(ws, r, "Budget", total_budget, MONEY_FMT)
    r = _kv(ws, r, "Actual", total_actual, MONEY_FMT)
    r = _kv(ws, r, "Variance", variance, MONEY_FMT)
    r = _kv(ws, r, "Variance %",
            variance / total_budget if total_budget else 0.0, PCT_FMT)
    if "impact" in detail.columns and len(detail):
        net_impact = float(detail["impact"].sum())
        r = _kv(ws, r, "P&L impact", net_impact, MONEY_FMT)
        r = _kv(ws, r, "Direction",
                "Favourable" if net_impact >= 0 else "Unfavourable")
    r = _kv(ws, r, "Flagged items", len(flagged))

    if len(flagged) and "pattern" in flagged.columns:
        r += 1
        ws.cell(row=r, column=1, value="Flagged by pattern").font = TITLE_FONT
        r += 1
        for pattern, count in flagged["pattern"].value_counts().items():
            r = _kv(ws, r, str(pattern), int(count))

    if len(detail):
        r += 1
        ws.cell(row=r, column=1, value="By department").font = TITLE_FONT
        r += 1
        dept = (
            detail.groupby("department", as_index=False)
            .agg(budget=("budget", "sum"), actual=("actual", "sum"),
                 impact=("impact", "sum"))
        )
        dept["variance"] = dept["actual"] - dept["budget"]
        dept["variance_pct"] = dept["variance"] / dept["budget"].abs()
        dept = dept.sort_values("impact")[
            ["department", "budget", "actual", "variance", "variance_pct", "impact"]
        ]
        _write_table(
            ws, dept, start_row=r,
            money_cols=("budget", "actual", "variance", "impact"),
            pct_cols=("variance_pct",),
            widths={"department": 22},
            highlight_col="impact",
        )
        ws.freeze_panes = None  # summary sheet reads better unfrozen

    # ---------------- Flagged ----------------
    ws2 = wb.create_sheet("Flagged Items")
    if len(flagged):
        cols = [c for c in [
            "month", "department", "account", "budget", "actual", "variance",
            "variance_pct", "severity", "pattern", "commentary",
        ] if c in flagged.columns]
        table = flagged.sort_values("impact")[cols] if "impact" in flagged.columns \
            else flagged[cols]
        _write_table(
            ws2, table,
            money_cols=("budget", "actual", "variance"),
            pct_cols=("variance_pct",),
            widths={"month": 10, "department": 18, "account": 26,
                    "severity": 10, "pattern": 18, "commentary": 90},
        )
    else:
        ws2["A1"] = "No items breached both thresholds for this selection."
        ws2["A1"].font = LABEL_FONT
        ws2.column_dimensions["A"].width = 60

    # ---------------- Detail ----------------
    ws3 = wb.create_sheet("Detail")
    if len(detail):
        cols = [c for c in [
            "month", "department", "account", "account_type", "budget",
            "actual", "variance", "variance_pct", "impact", "favorability",
            "ytd_budget", "ytd_actual", "ytd_variance", "ytd_variance_pct",
        ] if c in detail.columns]
        _write_table(
            ws3, detail.sort_values(["month", "department", "account"])[cols],
            money_cols=("budget", "actual", "variance", "impact",
                        "ytd_budget", "ytd_actual", "ytd_variance"),
            pct_cols=("variance_pct", "ytd_variance_pct"),
            widths={"month": 10, "department": 18, "account": 26},
        )
    else:
        ws3["A1"] = "No data for this selection."

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def template_csv_bytes() -> bytes:
    """A two-row starter CSV showing the schema an upload must match."""
    sample = pd.DataFrame(
        [
            ["2025-01", "Sales", "Revenue", "revenue", 1_600_000, 1_649_064],
            ["2025-01", "Marketing", "Advertising & Media", "expense", 160_000, 178_400],
        ],
        columns=TEMPLATE_COLUMNS,
    )
    return sample.to_csv(index=False).encode("utf-8")
