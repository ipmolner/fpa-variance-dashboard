"""
test_exports.py
---------------
Tests for exports.py. Run with:  pytest -q

The workbook is a deliverable someone may forward without opening it
first, so these check structure and number formats, not just that a file
was produced.
"""

import io

import openpyxl
import pandas as pd
import pytest

from commentary import add_commentary
from exports import TEMPLATE_COLUMNS, build_excel_report, template_csv_bytes
from variance_engine import build_variance_report, flag_variances, load_data, validate


@pytest.fixture(scope="module")
def report():
    return build_variance_report(load_data())


@pytest.fixture(scope="module")
def opex(report):
    return report[report["account_type"] == "expense"]


@pytest.fixture(scope="module")
def flagged(report, opex):
    return add_commentary(flag_variances(opex, dollar_floor=10_000), report)


def _wb(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data))


# --- structure -------------------------------------------------------------
def test_workbook_has_three_sheets(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    assert wb.sheetnames == ["Summary", "Flagged Items", "Detail"]


def test_flagged_sheet_row_count_matches(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    assert wb["Flagged Items"].max_row - 1 == len(flagged)


def test_detail_sheet_row_count_matches(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    assert wb["Detail"].max_row - 1 == len(opex)


def test_commentary_column_present_in_flagged(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    headers = [c.value for c in wb["Flagged Items"][1]]
    assert "Commentary" in headers


# --- summary content -------------------------------------------------------
def test_summary_totals_match_source(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    cells = {
        row[0]: row[1]
        for row in wb["Summary"].iter_rows(max_col=2, values_only=True)
        if row[0]
    }
    assert cells["Budget"] == pytest.approx(opex["budget"].sum())
    assert cells["Actual"] == pytest.approx(opex["actual"].sum())
    assert cells["Flagged items"] == len(flagged)


def test_summary_records_run_parameters(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged, "Operating expenses", 0.15, 25_000))
    cells = {
        row[0]: row[1]
        for row in wb["Summary"].iter_rows(max_col=2, values_only=True)
        if row[0]
    }
    assert cells["Variance threshold"] == pytest.approx(0.15)
    assert cells["Materiality floor"] == 25_000
    assert cells["View"] == "Operating expenses"


def test_direction_reflects_impact_sign(opex, flagged):
    wb = _wb(build_excel_report(opex, flagged))
    cells = {
        row[0]: row[1]
        for row in wb["Summary"].iter_rows(max_col=2, values_only=True)
        if row[0]
    }
    expected = "Favourable" if opex["impact"].sum() >= 0 else "Unfavourable"
    assert cells["Direction"] == expected


# --- formatting ------------------------------------------------------------
def test_money_columns_use_number_format(opex, flagged):
    ws = _wb(build_excel_report(opex, flagged))["Flagged Items"]
    headers = [c.value for c in ws[1]]
    budget_col = headers.index("Budget") + 1
    assert "#,##0" in ws.cell(row=2, column=budget_col).number_format


def test_pct_columns_use_percentage_format(opex, flagged):
    ws = _wb(build_excel_report(opex, flagged))["Flagged Items"]
    headers = [c.value for c in ws[1]]
    pct_col = headers.index("Variance Pct") + 1
    assert "0.0%" in ws.cell(row=2, column=pct_col).number_format


def test_flagged_sheet_has_frozen_header(opex, flagged):
    ws = _wb(build_excel_report(opex, flagged))["Flagged Items"]
    assert ws.freeze_panes == "A2"


# --- edge cases ------------------------------------------------------------
def test_empty_flagged_still_builds(opex):
    empty = opex.head(0).assign(pattern="", commentary="")
    wb = _wb(build_excel_report(opex, empty))
    assert wb["Flagged Items"]["A1"].value.startswith("No items breached")


def test_empty_detail_still_builds(opex, flagged):
    data = build_excel_report(opex.head(0), flagged.head(0))
    assert len(data) > 0
    assert _wb(data)["Detail"]["A1"].value == "No data for this selection."


def test_revenue_view_builds(report):
    rev = report[report["account_type"] == "revenue"]
    flags = add_commentary(flag_variances(rev, dollar_floor=10_000), report)
    wb = _wb(build_excel_report(rev, flags, "Revenue"))
    assert wb["Detail"].max_row - 1 == len(rev)


# --- template --------------------------------------------------------------
def test_template_has_required_columns():
    df = pd.read_csv(io.BytesIO(template_csv_bytes()))
    assert list(df.columns) == TEMPLATE_COLUMNS


def test_template_survives_validation():
    """The template must be a file the app would actually accept."""
    df = pd.read_csv(io.BytesIO(template_csv_bytes()))
    assert len(validate(df)) == 2
